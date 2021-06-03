import argparse
import csv
import json
from openpyxl import load_workbook

def run(excelfile, outputfile):

    wb = load_workbook(filename=excelfile)
    if wb.sheetnames.count('1_Testzahlerfassung') != 1:
        exit(1)
    wb.active = wb['1_Testzahlerfassung']
    ws = wb.active

    calendar_weeks = []
    weekly_tests = []
    parse_error = False

    for x in range(1, 10):
        calendar_weeks.append(f'2020-W0{x}')
        weekly_tests.append(0)

    row_index = 0
    for row in ws:
        if parse_error is True:
            break
        if row_index == 1:
            calendar_weeks.append(f'2020-W10')
            if str(row[2].value).isdecimal():
                weekly_tests.append(int(row[2].value))
            else:
                parse_error = True
                print('error: parsing first row.')
        elif row_index >= 2:
            if row[0].value == 'Summe' or row[0].value is None:
                break
            col_index = 0
            for col in row:
                if col_index == 0:
                    raw_calendar_week = str(col.value)
                    raw_week_array = raw_calendar_week.split('/')
                    raw_week = None
                    raw_year = None
                    if str(raw_week_array[0]).isdecimal() and str(raw_week_array[1]).isdecimal() and len(raw_week_array) == 2:
                        raw_week = int(raw_week_array[0])
                        raw_year = int(raw_week_array[1])
                    else:
                        parse_error = True
                        print('error: parsing raw_week_array.')
                        break
                    if raw_week is not None and raw_week > 0 and raw_week <= 53:
                        if raw_week < 10:
                            week = f'0{raw_week}'
                        else:
                            week = f'{raw_week}'
                    else:
                        parse_error = True
                        print('error: parsing raw_week.')
                        break
                    if raw_year is not None and raw_year >= 2020 and raw_year <= 9999:
                        calendar_week = f'{raw_year}-W{week}'
                        calendar_weeks.append(calendar_week)
                    else:
                        parse_error = True
                        print('error: parsing raw_year.')
                        break

                elif col_index == 1:
                    tests = None
                    if str(col.value).isdecimal():
                        tests = int(col.value)
                    else:
                        parse_error = True
                        print('error: parsing tests.')
                        break
                    if tests is not None and tests >= 0:
                        weekly_tests.append(tests)
                col_index +=1
        row_index +=1

    if parse_error is False:
        ''' data consistency check, length calendar_weeks equals length weekly_tests '''
        if len(calendar_weeks) != len(weekly_tests):
            print('length mismatch. ending programm.')
            print(f'calendar_weeks: {len(calendar_weeks)} weekly tests: {len(weekly_tests)}')
            exit(1)

        if len(calendar_weeks) < 1 and len(weekly_tests) < 1:
            print('no data extracted. ending programm.')
            exit(1)

        dict = {'calendar_weeks':calendar_weeks, 'weekly_tests':weekly_tests}
        with open(outputfile ,'w') as output:
            output.write(json.dumps(dict))

        print('wrote dict to file. program finished.')
        exit(0)
    else:
        print('error parsing weekly tests xslx.')
        exit(1)

if __name__ == '__main__':
    arg_parser = argparse.ArgumentParser()
    arg_parser.add_argument("-x", "--excelfile", type=str)
    arg_parser.add_argument("-o", "--outputfile", type=str)
    args = arg_parser.parse_args()
    if (args.excelfile and args.outputfile):
        run(args.excelfile, args.outputfile)
    else:
        arg_parser.print_help()
        exit(1)
