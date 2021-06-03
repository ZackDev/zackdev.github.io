import argparse
import csv
import json
import math
from openpyxl import load_workbook

def run(excelfile, outputfile):
    # open .xslx file, set working sheet to 'Impfungen_proTag'
    # select columns B C and D
    # clear direct formatting
    wb = load_workbook(filename=excelfile)
    if wb.sheetnames.count('Impfungen_proTag') != 1:
        exit(1)

    wb.active = wb['Impfungen_proTag']
    ws = wb.active

    dates = []
    primary_vaccinations = []
    secondary_vaccinations = []

    row_index = 0
    for row in ws:
        # first row is header row and doesn't contain usefull values
        if row_index > 0:
            col_index = 0
            row_value = str(row[0].value)
            # parse rows until end of interesting data
            if (( row_value is None ) or (row_value == '') or (row_value == 'None')):
                break
            for col in row:
                # the dates column
                if col_index == 0:
                    day, month, year = None, None, None
                    raw_date_str = str(col.value)
                    raw_date = raw_date_str.split(' ')[0]
                    raw_date_array = raw_date.split('-')
                    if len(raw_date_array) == 3:
                        if raw_date_array[0].isdigit() and raw_date_array[1].isdigit() and raw_date_array[2].isdigit():
                            year = int(raw_date_array[0])
                            month = int(raw_date_array[1])
                            day = int(raw_date_array[2])
                            if (2020 <= year) and (1 <= month <= 12) and (1 <= day <= 31):
                                dates.append(raw_date)
                            else:
                                print('error: day, month or year not in expected range.')
                                exit(1)
                        else:
                            print('error: day, month or year not numeric.')
                            exit(1)
                    else:
                        print('error: wrong date format.')
                        exit(1)

                # the primary vaccinations column
                elif col_index == 1:
                    raw_p_vacc = str(col.value)
                    if raw_p_vacc.isdigit():
                        p_vacc = int(raw_p_vacc)
                        if p_vacc >= 0:
                            primary_vaccinations.append(p_vacc)
                        else:
                            print('error: primary vaccinations negative.')
                            exit(1)
                    else:
                        print('error: primary vaccinations not numeric.')
                        exit(1)
                # the secondary vaccinations column
                elif col_index == 2:
                    raw_s_vacc = str(col.value)
                    if raw_s_vacc.isdigit():
                        s_vacc = int(raw_s_vacc)
                        if s_vacc >= 0:
                            secondary_vaccinations.append(s_vacc)
                        else:
                            print('error: secondary vaccinations not numeric.')
                            exit(1)
                    elif raw_s_vacc is None or raw_s_vacc == 'None':
                        secondary_vaccinations.append(int(0))
                    else:
                        print('error: secondary vaccinations not numeric nor default zero.')
                        exit(1)
                col_index +=1
        row_index +=1

    # write csvfile
    ''' check data for consistency, equal amount of dates and cases '''
    if len(dates) != len(primary_vaccinations) != len(secondary_vaccinations):
        print('length mismatch. ending programm.')
        print(f'dates: {len(dates)} primary_vaccinations: {len(primary_vaccinations)} secondary_vaccinations: {len(secondary_vaccinations)}')
        exit(1)

    if len(dates) < 1 and len(primary_vaccinations) < 1 and len(secondary_vaccinations):
        print('no data extracted. ending programm.')
        exit(1)

    dict = { 'dates':dates, 'primary_vaccinations':primary_vaccinations, 'secondary_vaccinations':secondary_vaccinations }
    with open(outputfile ,'w') as output:
        output.write(json.dumps(dict))

    print('wrote dict to file. program finished.')
    exit(0)

if __name__ == '__main__':
    arg_parser = argparse.ArgumentParser()
    arg_parser.add_argument("-x", "--excelfile", type=str)
    arg_parser.add_argument("-o", "--outputfile", type=str)
    args = arg_parser.parse_args()
    if (args.excelfile and args.outputfile):
        run(args.excelfile, args.outputfile)
    else:
        arg_parser.print_help()
        exit(1)
